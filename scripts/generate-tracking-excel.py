"""
Generate TrafficMENA Events Tracking Data Model Excel file.
Mirrors the structure of the reference AlMujtama Pharmacy data model
but tailored for TrafficMENA education platform.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ── Style definitions ──────────────────────────────────────────────

HEADER_FONT = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")

EVENT_FONT = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
EVENT_FILL = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")

COL_HEADER_FONT = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
COL_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

HIGH_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
MEDIUM_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
LOW_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

PII_FILL = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")

PARAM_FONT = Font(name="Calibri", size=10)
CODE_FONT = Font(name="Consolas", size=9)
WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")
TOP_ALIGN = Alignment(vertical="top")

THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


def style_row(ws, row, font=PARAM_FONT, fill=None, align=TOP_ALIGN, border=THIN_BORDER):
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = font
        cell.alignment = align
        cell.border = border
        if fill:
            cell.fill = fill


def priority_fill(priority):
    if priority == "High":
        return HIGH_FILL
    if priority == "Medium":
        return MEDIUM_FILL
    return LOW_FILL


def write_event_header(ws, row, event_name, trigger):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    cell = ws.cell(row=row, column=1, value=f"Event name : {event_name}      |      trigger : {trigger}")
    cell.font = EVENT_FONT
    cell.fill = EVENT_FILL
    cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[row].height = 28
    for c in range(2, 8):
        ws.cell(row=row, column=c).fill = EVENT_FILL
    return row + 1


def write_param(ws, row, name, ptype, sample, notes, json_sample, priority, status, is_pii=False):
    ws.cell(row=row, column=1, value=name)
    ws.cell(row=row, column=2, value=ptype)
    ws.cell(row=row, column=3, value=sample)
    ws.cell(row=row, column=4, value=notes)
    if json_sample:
        cell = ws.cell(row=row, column=5, value=json_sample)
        cell.font = CODE_FONT
        cell.alignment = WRAP_ALIGN
    ws.cell(row=row, column=6, value=priority)
    ws.cell(row=row, column=7, value=status)

    style_row(ws, row)
    ws.cell(row=row, column=6).fill = priority_fill(priority)

    if is_pii:
        ws.cell(row=row, column=1).fill = PII_FILL
        ws.cell(row=row, column=1).font = Font(name="Calibri", size=10, bold=True)

    return row + 1


# ════════════════════════════════════════════════════════════════════
# SHEET 1: Event Taxonomy
# ════════════════════════════════════════════════════════════════════

ws_tax = wb.active
ws_tax.title = "Event Taxonomy"
ws_tax.sheet_properties.tabColor = "1F4E79"

# Title
ws_tax.merge_cells("A1:F1")
ws_tax.cell(row=1, column=1, value="TRAFFICMENA — EVENTS TRACKING DATA MODEL").font = Font(name="Calibri", bold=True, size=16, color="FFFFFF")
ws_tax.cell(row=1, column=1).fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
for c in range(2, 7):
    ws_tax.cell(row=1, column=c).fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")

ws_tax.merge_cells("A2:F2")
ws_tax.cell(row=2, column=1, value="Platform: GTM → GA4  |  Scope: Public + Member/Expert Dashboard  |  Currency: Dynamic (EGP)  |  Version: 1.0  |  Date: 2026-04-14").font = Font(name="Calibri", size=10, italic=True)

# Column headers
headers = ["#", "Event Name", "Trigger", "GA4 Standard", "Priority", "Tracking"]
for i, h in enumerate(headers, 1):
    cell = ws_tax.cell(row=4, column=i, value=h)
    cell.font = COL_HEADER_FONT
    cell.fill = COL_HEADER_FILL
    cell.alignment = Alignment(horizontal="center")

events = [
    (1, "global_variables", "Every page load", "Custom", "High", "Client"),
    (2, "login_start", "User submits email for OTP", "Custom", "High", "Client"),
    (3, "login", "OTP verified successfully", "login (GA4)", "High", "Client"),
    (4, "sign_up_step", "Each signup wizard step completed", "Custom", "High", "Client"),
    (5, "sign_up", "Signup wizard fully completed", "sign_up (GA4)", "High", "Client + Server"),
    (6, "view_item_list", "Events or tracks listing page viewed", "view_item_list (GA4)", "High", "Client"),
    (7, "select_item", "User clicks event/track card from list", "select_item (GA4)", "Medium", "Client"),
    (8, "view_item", "Event or track detail page viewed", "view_item (GA4)", "High", "Client"),
    (9, "event_registration", "User registers for a free event", "Custom", "High", "Client + Server"),
    (10, "track_booking", "User books a free track", "Custom", "High", "Client + Server"),
    (11, "begin_checkout", "Payment dialog opens", "begin_checkout (GA4)", "High", "Client"),
    (12, "select_payment_method", "User clicks Pay after selecting payment method", "Custom", "High", "Client"),
    (13, "apply_promo_code", "Promo code submitted", "Custom", "Medium", "Client"),
    (14, "first_purchase", "First-ever non-subscription purchase", "Custom", "High", "Client + Server"),
    (15, "purchase", "Every non-subscription purchase (including first)", "purchase (GA4)", "High", "Client + Server"),
    (16, "subscribe", "Subscription purchase (new or returning)", "Custom", "High", "Client + Server"),
    (17, "click_meeting_link", "User clicks Zoom/meeting link to attend", "Custom", "High", "Client"),
    (18, "view_content", "User opens a lesson or library asset", "Custom", "High", "Client"),
    (19, "download_content", "User downloads a library asset", "Custom", "Medium", "Client"),
    (20, "add_to_calendar", "User adds event to calendar", "Custom", "Medium", "Client"),
    (21, "calculator_used", "User completes a calculator calculation", "Custom", "Medium", "Client"),
    (22, "cancel_registration", "User cancels registration or requests refund", "Custom", "High", "Client + Server"),
    (23, "profile_updated", "User updates profile information", "Custom", "Low", "Client"),
]

for i, ev in enumerate(events):
    row = 5 + i
    for j, val in enumerate(ev):
        cell = ws_tax.cell(row=row, column=j + 1, value=val)
        cell.font = PARAM_FONT
        cell.border = THIN_BORDER
        cell.alignment = TOP_ALIGN
    ws_tax.cell(row=row, column=5).fill = priority_fill(ev[4])

ws_tax.column_dimensions["A"].width = 5
ws_tax.column_dimensions["B"].width = 22
ws_tax.column_dimensions["C"].width = 45
ws_tax.column_dimensions["D"].width = 25
ws_tax.column_dimensions["E"].width = 12
ws_tax.column_dimensions["F"].width = 18


# ════════════════════════════════════════════════════════════════════
# SHEET 2: Data Model (main sheet — mirrors the reference)
# ════════════════════════════════════════════════════════════════════

ws = wb.create_sheet("Data Model")
ws.sheet_properties.tabColor = "2E75B6"

# Title rows
ws.merge_cells("A1:G1")
ws.cell(row=1, column=1, value="DATA MODEL  |  TrafficMENA  |  Events Tracking Data Model").font = HEADER_FONT
ws.cell(row=1, column=1).fill = HEADER_FILL
for c in range(2, 8):
    ws.cell(row=1, column=c).fill = HEADER_FILL

ws.merge_cells("A2:G2")
ws.cell(row=2, column=1, value="Platform: GTM → GA4  |  Currency: Dynamic (EGP)  |  PII fields highlighted in orange — block from GA4").font = Font(name="Calibri", size=10, italic=True)

# Column headers
col_headers = ["Parameter Name", "Parameter Type", "Parameter Sample", "Notes", "JSON Sample", "Priority Level", "Status"]
for i, h in enumerate(col_headers, 1):
    cell = ws.cell(row=3, column=i, value=h)
    cell.font = COL_HEADER_FONT
    cell.fill = COL_HEADER_FILL
    cell.alignment = Alignment(horizontal="center")

ws.column_dimensions["A"].width = 24
ws.column_dimensions["B"].width = 14
ws.column_dimensions["C"].width = 38
ws.column_dimensions["D"].width = 55
ws.column_dimensions["E"].width = 65
ws.column_dimensions["F"].width = 16
ws.column_dimensions["G"].width = 16

row = 4

# ── 1. global_variables ──

GLOBAL_JSON = """window.dataLayer = window.dataLayer || [];

window.dataLayer.push({
  'event': 'global_variables',
  'user_id': 'a1b2c3d4-e5f6-7890-abcd-ef1234567890',
  'login_status': 'logged_in',
  'user_role': 'user',
  'subscription_status': 'active',
  'customer_type': 'returning',
  'event_source': 'Web',
  'page_type': 'event_list',
  'currency': 'EGP',
  'total_registrations': 3,
  'total_purchases': 2,
  'total_revenue': 750.00,
  'account_creation_date': '2026-01-15T10:30:00Z',
  'email': 'ahmed@example.com',
  'phone': '201012345678',
  'first_name': 'Ahmed',
  'last_name': 'Hassan'
});"""

row = write_event_header(ws, row, "global_variables", "on every page load")
row = write_param(ws, row, "user_id", "String", "a1b2c3d4-e5f6-...", "UUID from users.id. Empty string for guests.", GLOBAL_JSON, "High", "To Be Added")
row = write_param(ws, row, "login_status", "String", "logged_in", '"logged_in" or "guest"', "", "High", "To Be Added")
row = write_param(ws, row, "user_role", "String", "user", '"user" / "expert" — member dashboard roles only', "", "Medium", "To Be Added")
row = write_param(ws, row, "subscription_status", "String", "none", '"active" / "expired" / "none"', "", "High", "To Be Added")
row = write_param(ws, row, "customer_type", "String", "returning", '"free" (0 purchases) / "new" (1 purchase) / "returning" (2+). Derived from total_purchases.', "", "High", "To Be Added")
row = write_param(ws, row, "event_source", "String", "Web", 'Platform. Future: "iOS", "Android"', "", "High", "To Be Added")
row = write_param(ws, row, "page_type", "String", "homepage", "Page type identifier (see Page Types sheet)", "", "Medium", "To Be Added")
row = write_param(ws, row, "currency", "String", "EGP", "Dynamic from platform settings. Multi-currency planned.", "", "High", "To Be Added")
row = write_param(ws, row, "total_registrations", "Number", "3", "Total events the user has registered for", "", "Medium", "To Be Added")
row = write_param(ws, row, "total_purchases", "Number", "2", "Total successful purchases", "", "Medium", "To Be Added")
row = write_param(ws, row, "total_revenue", "Number", "750.00", "Total amount spent (currency units, not cents)", "", "Medium", "To Be Added")
row = write_param(ws, row, "account_creation_date", "Date", "2026-01-15T10:30:00Z", "ISO-8601 format", "", "Medium", "To Be Added")
row = write_param(ws, row, "email", "String", "ahmed@example.com", "PII — block from GA4. Include for CRM/marketing tags.", "", "High", "To Be Added", is_pii=True)
row = write_param(ws, row, "phone", "String", "201012345678", "PII — digits only, no + prefix. Block from GA4.", "", "High", "To Be Added", is_pii=True)
row = write_param(ws, row, "first_name", "String", "Ahmed", "PII — block from GA4.", "", "High", "To Be Added", is_pii=True)
row = write_param(ws, row, "last_name", "String", "Hassan", "PII — block from GA4.", "", "High", "To Be Added", is_pii=True)
row += 1

# ── 2. login_start ──

LOGIN_START_JSON = """window.dataLayer.push({
  'event': 'login_start',
  'method': 'email_otp',
  'event_source': 'Web'
});"""

row = write_event_header(ws, row, "login_start", "user submits email on /signin to request OTP")
row = write_param(ws, row, "method", "String", "email_otp", "Authentication method used", LOGIN_START_JSON, "High", "To Be Added")
row = write_param(ws, row, "event_source", "String", "Web", "Platform identifier", "", "High", "To Be Added")
row += 1

# ── 3. login ──

LOGIN_JSON = """window.dataLayer.push({
  'event': 'login',
  'method': 'email_otp',
  'status': 'success',
  'user_id': 'a1b2c3d4-e5f6-7890-abcd-ef1234567890',
  'event_source': 'Web',
  'email': 'ahmed@example.com'
});"""

row = write_event_header(ws, row, "login", "OTP verified successfully, session established")
row = write_param(ws, row, "method", "String", "email_otp", "Authentication method", LOGIN_JSON, "High", "To Be Added")
row = write_param(ws, row, "status", "String", "success", '"success" or "failure"', "", "High", "To Be Added")
row = write_param(ws, row, "user_id", "String", "a1b2c3d4-...", "UUID assigned after verification", "", "High", "To Be Added")
row = write_param(ws, row, "event_source", "String", "Web", "Platform", "", "High", "To Be Added")
row = write_param(ws, row, "email", "String", "ahmed@example.com", "PII — block from GA4", "", "High", "To Be Added", is_pii=True)
row += 1

# ── 4. sign_up_step ──

SIGNUP_STEP_JSON = """// Step 1 — name entered
window.dataLayer.push({
  'event': 'sign_up_step',
  'step_number': 1,
  'step_name': 'name_info',
  'event_source': 'Web'
});

// Step 3 — OTP verified
window.dataLayer.push({
  'event': 'sign_up_step',
  'step_number': 3,
  'step_name': 'otp_verified',
  'event_source': 'Web'
});"""

row = write_event_header(ws, row, "sign_up_step", "each signup wizard step completed (steps 1-6)")
row = write_param(ws, row, "step_number", "Number", "1", "Sequential step: 1=name_info, 2=email_entered, 3=otp_verified, 4=phone_entered, 5=goal_selected, 6=challenge_selected", SIGNUP_STEP_JSON, "High", "To Be Added")
row = write_param(ws, row, "step_name", "String", "name_info", "Human-readable step ID. See step definitions in notes.", "", "High", "To Be Added")
row = write_param(ws, row, "event_source", "String", "Web", "Platform", "", "High", "To Be Added")
row += 1

# ── 5. sign_up ──

SIGNUP_JSON = """window.dataLayer.push({
  'event': 'sign_up',
  'method': 'email_otp',
  'user_id': 'a1b2c3d4-e5f6-7890-abcd-ef1234567890',
  'event_source': 'Web',
  'account_creation_date': '2026-04-14T14:30:00Z',
  'email': 'ahmed@example.com',
  'phone': '201012345678',
  'first_name': 'Ahmed',
  'last_name': 'Hassan'
});"""

row = write_event_header(ws, row, "sign_up", "signup wizard fully completed — account created (Client + Server)")
row = write_param(ws, row, "method", "String", "email_otp", "Signup method", SIGNUP_JSON, "High", "To Be Added")
row = write_param(ws, row, "user_id", "String", "a1b2c3d4-...", "Newly created user UUID", "", "High", "To Be Added")
row = write_param(ws, row, "event_source", "String", "Web", "Platform. Server-side sends 'Server'", "", "High", "To Be Added")
row = write_param(ws, row, "account_creation_date", "Date", "2026-04-14T14:30:00Z", "ISO-8601", "", "Medium", "To Be Added")
row = write_param(ws, row, "email", "String", "ahmed@example.com", "PII — block from GA4", "", "High", "To Be Added", is_pii=True)
row = write_param(ws, row, "phone", "String", "201012345678", "PII — digits only, no +", "", "High", "To Be Added", is_pii=True)
row = write_param(ws, row, "first_name", "String", "Ahmed", "PII — block from GA4", "", "High", "To Be Added", is_pii=True)
row = write_param(ws, row, "last_name", "String", "Hassan", "PII — block from GA4", "", "High", "To Be Added", is_pii=True)
row += 1

# ── 6. view_item_list ──

VIEW_LIST_JSON = """window.dataLayer.push({
  'event': 'view_item_list',
  'item_list_id': 'events',
  'item_list_name': 'Events',
  'items': [
    {
      'item_id': 'evt-a1b2c3d4-e5f6-...',
      'item_name': 'Advanced Google Ads Workshop',
      'item_category': 'Mastermind',
      'price': 250.00,
      'currency': 'EGP',
      'item_image_link': 'https://trafficmena.b-cdn.net/events/workshop.jpg',
      'item_link': '/meetups/evt-a1b2c3d4-...',
      'index': 0
    },
    {
      'item_id': 'evt-b2c3d4e5-f6a7-...',
      'item_name': 'SEO Fundamentals Meetup',
      'item_category': 'Meetup',
      'price': 0,
      'currency': 'EGP',
      'item_image_link': 'https://trafficmena.b-cdn.net/events/seo.jpg',
      'item_link': '/meetups/evt-b2c3d4e5-...',
      'index': 1
    }
  ]
});"""

row = write_event_header(ws, row, "view_item_list", "events or tracks listing page viewed (/meetups or tracks page)")
row = write_param(ws, row, "item_list_id", "String", "events", '"events" or "tracks"', VIEW_LIST_JSON, "High", "To Be Added")
row = write_param(ws, row, "item_list_name", "String", "Events", '"Events" or "Tracks"', "", "High", "To Be Added")
row = write_param(ws, row, "items[].item_id", "String", "evt-a1b2c3d4-...", "Event/track UUID", "", "High", "To Be Added")
row = write_param(ws, row, "items[].item_name", "String", "Advanced Google Ads Workshop", "Event/track title", "", "High", "To Be Added")
row = write_param(ws, row, "items[].item_category", "String", "Mastermind", 'Event type: Event/Meetup/Mastermind/Retreat. For tracks: "Track"', "", "High", "To Be Added")
row = write_param(ws, row, "items[].price", "Number", "250.00", "Price in currency units (priceInCents / 100). 0 for free.", "", "High", "To Be Added")
row = write_param(ws, row, "items[].currency", "String", "EGP", "Dynamic currency", "", "High", "To Be Added")
row = write_param(ws, row, "items[].item_image_link", "String", "https://trafficmena.b-cdn.net/...", "CDN image URL", "", "Low", "To Be Added")
row = write_param(ws, row, "items[].item_link", "String", "/meetups/evt-a1b2c3d4-...", "Relative detail page URL", "", "Medium", "To Be Added")
row = write_param(ws, row, "items[].index", "Number", "0", "Position in the list (0-based)", "", "Medium", "To Be Added")
row += 1

# ── 7. select_item ──

SELECT_JSON = """window.dataLayer.push({
  'event': 'select_item',
  'item_list_id': 'events',
  'item_list_name': 'Events',
  'items': [
    {
      'item_id': 'evt-a1b2c3d4-...',
      'item_name': 'Advanced Google Ads Workshop',
      'item_category': 'Mastermind',
      'price': 250.00,
      'currency': 'EGP',
      'index': 0
    }
  ]
});"""

row = write_event_header(ws, row, "select_item", "user clicks on an event or track card from listing page")
row = write_param(ws, row, "item_list_id", "String", "events", "Which list the item was clicked from", SELECT_JSON, "Medium", "To Be Added")
row = write_param(ws, row, "item_list_name", "String", "Events", "Display name of the list", "", "Medium", "To Be Added")
row = write_param(ws, row, "items[]", "Array", "(single item)", "Same item parameters as view_item_list", "", "Medium", "To Be Added")
row += 1

# ── 8. view_item ──

VIEW_ITEM_JSON = """// Paid offline event
window.dataLayer.push({
  'event': 'view_item',
  'currency': 'EGP',
  'value': 250.00,
  'items': [
    {
      'item_id': 'evt-a1b2c3d4-...',
      'item_name': 'Advanced Google Ads Workshop',
      'item_category': 'Mastermind',
      'price': 250.00,
      'currency': 'EGP',
      'item_image_link': 'https://trafficmena.b-cdn.net/events/workshop.jpg',
      'item_link': '/meetups/evt-a1b2c3d4-...',
      'item_location': 'Cairo, Egypt',
      'item_date': '2026-05-20T18:00:00Z',
      'is_online': false,
      'spots_remaining': 12
    }
  ]
});

// Track detail page
window.dataLayer.push({
  'event': 'view_item',
  'currency': 'EGP',
  'value': 1500.00,
  'items': [
    {
      'item_id': 'trk-c3d4e5f6-...',
      'item_name': 'Performance Marketing Bootcamp',
      'item_category': 'Track',
      'price': 1500.00,
      'currency': 'EGP',
      'item_location': 'Cairo, Egypt',
      'is_online': false,
      'spots_remaining': 8
    }
  ]
});"""

row = write_event_header(ws, row, "view_item", "event detail (/meetups/:id) or track detail (/tracks/:id) page loads")
row = write_param(ws, row, "currency", "String", "EGP", "Dynamic currency", VIEW_ITEM_JSON, "High", "To Be Added")
row = write_param(ws, row, "value", "Number", "250.00", "Item price (priceInCents / 100)", "", "High", "To Be Added")
row = write_param(ws, row, "items[].item_id", "String", "evt-a1b2c3d4-...", "Event/track UUID", "", "High", "To Be Added")
row = write_param(ws, row, "items[].item_name", "String", "Advanced Google Ads Workshop", "Title", "", "High", "To Be Added")
row = write_param(ws, row, "items[].item_category", "String", "Mastermind", 'Event type or "Track"', "", "High", "To Be Added")
row = write_param(ws, row, "items[].price", "Number", "250.00", "Price (priceInCents / 100)", "", "High", "To Be Added")
row = write_param(ws, row, "items[].currency", "String", "EGP", "Dynamic", "", "High", "To Be Added")
row = write_param(ws, row, "items[].item_image_link", "String", "https://trafficmena.b-cdn.net/...", "CDN image", "", "Medium", "To Be Added")
row = write_param(ws, row, "items[].item_link", "String", "/meetups/evt-a1b2c3d4-...", "Page URL", "", "Medium", "To Be Added")
row = write_param(ws, row, "items[].item_location", "String", "Cairo, Egypt", "Physical location. Empty for online events.", "", "Medium", "To Be Added")
row = write_param(ws, row, "items[].item_date", "Date", "2026-05-20T18:00:00Z", "Event date ISO-8601", "", "Medium", "To Be Added")
row = write_param(ws, row, "items[].is_online", "Boolean", "false", "true if event has meetingLink and no location", "", "Medium", "To Be Added")
row = write_param(ws, row, "items[].spots_remaining", "Number", "12", "Remaining capacity. null if unlimited.", "", "Low", "To Be Added")
row += 1

# ── 9. event_registration ──

EVENT_REG_JSON = """window.dataLayer.push({
  'event': 'event_registration',
  'item_id': 'evt-b2c3d4e5-...',
  'item_name': 'SEO Fundamentals Meetup',
  'item_category': 'Meetup',
  'registration_type': 'free',
  'is_online': true,
  'event_source': 'Web'
});"""

row = write_event_header(ws, row, "event_registration", "user registers for a FREE event (Client + Server). Do NOT fire for paid events.")
row = write_param(ws, row, "item_id", "String", "evt-b2c3d4e5-...", "Event UUID", EVENT_REG_JSON, "High", "To Be Added")
row = write_param(ws, row, "item_name", "String", "SEO Fundamentals Meetup", "Event title", "", "High", "To Be Added")
row = write_param(ws, row, "item_category", "String", "Meetup", "Event type", "", "High", "To Be Added")
row = write_param(ws, row, "registration_type", "String", "free", 'Always "free" for this event', "", "High", "To Be Added")
row = write_param(ws, row, "is_online", "Boolean", "true", "Online vs offline", "", "Medium", "To Be Added")
row = write_param(ws, row, "event_source", "String", "Web", 'Platform. Server sends "Server"', "", "High", "To Be Added")
row += 1

# ── 10. track_booking ──

TRACK_BOOK_JSON = """window.dataLayer.push({
  'event': 'track_booking',
  'item_id': 'trk-d4e5f6a7-...',
  'item_name': 'Content Marketing Fundamentals',
  'item_category': 'Track',
  'booking_type': 'free',
  'event_count': 5,
  'event_source': 'Web'
});"""

row = write_event_header(ws, row, "track_booking", "user books a FREE track (Client + Server). Do NOT fire for paid tracks.")
row = write_param(ws, row, "item_id", "String", "trk-d4e5f6a7-...", "Track UUID", TRACK_BOOK_JSON, "High", "To Be Added")
row = write_param(ws, row, "item_name", "String", "Content Marketing Fundamentals", "Track title", "", "High", "To Be Added")
row = write_param(ws, row, "item_category", "String", "Track", 'Always "Track"', "", "High", "To Be Added")
row = write_param(ws, row, "booking_type", "String", "free", 'Always "free"', "", "High", "To Be Added")
row = write_param(ws, row, "event_count", "Number", "5", "Number of events in the track", "", "Medium", "To Be Added")
row = write_param(ws, row, "event_source", "String", "Web", "Platform", "", "High", "To Be Added")
row += 1

# ── 11. begin_checkout ──

BEGIN_CHECKOUT_JSON = """// Paid event checkout
window.dataLayer.push({
  'event': 'begin_checkout',
  'currency': 'EGP',
  'value': 250.00,
  'item_type': 'event_ticket',
  'items': [{
    'item_id': 'evt-a1b2c3d4-...',
    'item_name': 'Advanced Google Ads Workshop',
    'item_category': 'Mastermind',
    'price': 250.00,
    'currency': 'EGP'
  }]
});

// Subscription checkout
window.dataLayer.push({
  'event': 'begin_checkout',
  'currency': 'EGP',
  'value': 2500.00,
  'item_type': 'subscription',
  'items': [{
    'item_id': 'subscription_annual',
    'item_name': 'TrafficMENA Annual Subscription',
    'item_category': 'Subscription',
    'price': 2500.00,
    'currency': 'EGP'
  }]
});"""

row = write_event_header(ws, row, "begin_checkout", "payment dialog opens for paid event, track, or subscription")
row = write_param(ws, row, "currency", "String", "EGP", "Dynamic", BEGIN_CHECKOUT_JSON, "High", "To Be Added")
row = write_param(ws, row, "value", "Number", "250.00", "Total price before discount (cents / 100)", "", "High", "To Be Added")
row = write_param(ws, row, "item_type", "String", "event_ticket", '"event_ticket" / "track_booking" / "subscription"', "", "High", "To Be Added")
row = write_param(ws, row, "items[]", "Array", "(single item)", "item_id, item_name, item_category, price, currency", "", "High", "To Be Added")
row += 1

# ── 12. select_payment_method ──

SELECT_PAYMENT_JSON = """// Fires when user clicks "Pay" in PaymentCheckoutDialog
// NOT when they select the radio button
window.dataLayer.push({
  'event': 'select_payment_method',
  'currency': 'EGP',
  'value': 200.00,
  'payment_type': 'fawry',
  'item_type': 'event_ticket',
  'coupon': 'SUMMER25',
  'items': [{
    'item_id': 'evt-a1b2c3d4-...',
    'item_name': 'Advanced Google Ads Workshop',
    'item_category': 'Mastermind',
    'price': 200.00,
    'currency': 'EGP'
  }]
});"""

row = write_event_header(ws, row, "select_payment_method", 'user clicks "Pay" after selecting method. Actual payment info entered on Fawaterk.')
row = write_param(ws, row, "currency", "String", "EGP", "Dynamic", SELECT_PAYMENT_JSON, "High", "To Be Added")
row = write_param(ws, row, "value", "Number", "200.00", "Amount after discount", "", "High", "To Be Added")
row = write_param(ws, row, "payment_type", "String", "fawry", "Selected method from Fawaterk API (fawry/aman/masary/meeza/mobile_wallet)", "", "High", "To Be Added")
row = write_param(ws, row, "item_type", "String", "event_ticket", '"event_ticket" / "track_booking" / "subscription"', "", "High", "To Be Added")
row = write_param(ws, row, "coupon", "String", "SUMMER25", "Promo code if applied. Empty string if none.", "", "Medium", "To Be Added")
row = write_param(ws, row, "items[]", "Array", "(single item)", "item_id, item_name, item_category, price, currency", "", "High", "To Be Added")
row += 1

# ── 13. apply_promo_code ──

PROMO_JSON = """// Successful promo code
window.dataLayer.push({
  'event': 'apply_promo_code',
  'promo_code': 'SUMMER25',
  'status': 'success',
  'discount_percent': 25,
  'item_type': 'event_ticket',
  'item_id': 'evt-a1b2c3d4-...'
});

// Failed promo code
window.dataLayer.push({
  'event': 'apply_promo_code',
  'promo_code': 'OLDCODE',
  'status': 'expired',
  'discount_percent': 0,
  'item_type': 'event_ticket',
  'item_id': 'evt-a1b2c3d4-...'
});"""

row = write_event_header(ws, row, "apply_promo_code", "user submits a promo code (success or failure)")
row = write_param(ws, row, "promo_code", "String", "SUMMER25", "The code entered", PROMO_JSON, "High", "To Be Added")
row = write_param(ws, row, "status", "String", "success", '"success" / "invalid" / "expired" / "limit_reached"', "", "High", "To Be Added")
row = write_param(ws, row, "discount_percent", "Number", "25", "Discount percentage. 0 if failed.", "", "Medium", "To Be Added")
row = write_param(ws, row, "item_type", "String", "event_ticket", "What the code was applied to", "", "Medium", "To Be Added")
row = write_param(ws, row, "item_id", "String", "evt-a1b2c3d4-...", "Target item UUID", "", "Medium", "To Be Added")
row += 1

# ── 14. first_purchase ──

FIRST_PURCHASE_JSON = """// Fires ADDITIONALLY alongside purchase
// when user has 0 prior non-subscription purchases
window.dataLayer.push({
  'event': 'first_purchase',
  'transaction_id': 'pay-e5f6a7b8-c9d0-...',
  'currency': 'EGP',
  'value': 200.00,
  'item_type': 'event_ticket',
  'payment_type': 'fawry',
  'customer_type': 'new',
  'coupon': 'SUMMER25',
  'discount': 50.00,
  'original_value': 250.00,
  'items': [{
    'item_id': 'evt-a1b2c3d4-...',
    'item_name': 'Advanced Google Ads Workshop',
    'item_category': 'Mastermind',
    'price': 200.00,
    'currency': 'EGP',
    'quantity': 1
  }]
});"""

row = write_event_header(ws, row, "first_purchase", "fires ADDITIONALLY alongside purchase on first-ever non-subscription purchase (Client + Server)")
row = write_param(ws, row, "transaction_id", "String", "pay-e5f6a7b8-...", "Payment UUID from payments.id", FIRST_PURCHASE_JSON, "High", "To Be Added")
row = write_param(ws, row, "currency", "String", "EGP", "From payments.currency", "", "High", "To Be Added")
row = write_param(ws, row, "value", "Number", "200.00", "Amount paid (amountCents / 100)", "", "High", "To Be Added")
row = write_param(ws, row, "item_type", "String", "event_ticket", '"event_ticket" / "track_booking"', "", "High", "To Be Added")
row = write_param(ws, row, "payment_type", "String", "fawry", "Payment method used", "", "High", "To Be Added")
row = write_param(ws, row, "customer_type", "String", "new", 'Always "new" on first_purchase — member just became a customer', "", "High", "To Be Added")
row = write_param(ws, row, "coupon", "String", "SUMMER25", "Promo code if used. Empty string if none.", "", "Medium", "To Be Added")
row = write_param(ws, row, "discount", "Number", "50.00", "Discount amount (discountAppliedCents / 100). 0 if none.", "", "Medium", "To Be Added")
row = write_param(ws, row, "original_value", "Number", "250.00", "Pre-discount price (originalAmountCents / 100)", "", "Medium", "To Be Added")
row = write_param(ws, row, "items[].item_id", "String", "evt-a1b2c3d4-...", "Event/track UUID", "", "High", "To Be Added")
row = write_param(ws, row, "items[].item_name", "String", "Advanced Google Ads Workshop", "Title", "", "High", "To Be Added")
row = write_param(ws, row, "items[].item_category", "String", "Mastermind", 'Event type or "Track"', "", "High", "To Be Added")
row = write_param(ws, row, "items[].price", "Number", "200.00", "Price paid", "", "High", "To Be Added")
row = write_param(ws, row, "items[].currency", "String", "EGP", "Dynamic", "", "High", "To Be Added")
row = write_param(ws, row, "items[].quantity", "Number", "1", "Always 1 (single registration)", "", "High", "To Be Added")
row += 1

# ── 15. purchase ──

PURCHASE_JSON = """// First-ever purchase — fires BOTH purchase AND first_purchase
window.dataLayer.push({
  'event': 'purchase',
  'transaction_id': 'pay-e5f6a7b8-c9d0-...',
  'currency': 'EGP',
  'value': 200.00,
  'item_type': 'event_ticket',
  'payment_type': 'fawry',
  'customer_type': 'new',
  'coupon': 'SUMMER25',
  'discount': 50.00,
  'original_value': 250.00,
  'items': [{
    'item_id': 'evt-a1b2c3d4-...',
    'item_name': 'Advanced Google Ads Workshop',
    'item_category': 'Mastermind',
    'price': 200.00,
    'currency': 'EGP',
    'quantity': 1
  }]
});

// Returning customer — fires purchase only
window.dataLayer.push({
  'event': 'purchase',
  'transaction_id': 'pay-f6a7b8c9-d0e1-...',
  'currency': 'EGP',
  'value': 1500.00,
  'item_type': 'track_booking',
  'payment_type': 'aman',
  'customer_type': 'returning',
  'coupon': '',
  'discount': 0,
  'original_value': 1500.00,
  'items': [{
    'item_id': 'trk-c3d4e5f6-...',
    'item_name': 'Performance Marketing Bootcamp',
    'item_category': 'Track',
    'price': 1500.00,
    'currency': 'EGP',
    'quantity': 1
  }]
});"""

row = write_event_header(ws, row, "purchase", "EVERY non-subscription purchase (Client + Server). On first purchase, also fires first_purchase alongside.")
row = write_param(ws, row, "transaction_id", "String", "pay-f6a7b8c9-...", "Payment UUID", PURCHASE_JSON, "High", "To Be Added")
row = write_param(ws, row, "currency", "String", "EGP", "Dynamic", "", "High", "To Be Added")
row = write_param(ws, row, "value", "Number", "1500.00", "Amount paid", "", "High", "To Be Added")
row = write_param(ws, row, "item_type", "String", "track_booking", '"event_ticket" / "track_booking"', "", "High", "To Be Added")
row = write_param(ws, row, "payment_type", "String", "aman", "Payment method", "", "High", "To Be Added")
row = write_param(ws, row, "customer_type", "String", "returning", '"new" (first purchase, 0 prior) / "returning" (1+ prior purchases)', "", "High", "To Be Added")
row = write_param(ws, row, "coupon", "String", "", "Promo code. Empty if none.", "", "Medium", "To Be Added")
row = write_param(ws, row, "discount", "Number", "0", "Discount amount. 0 if none.", "", "Medium", "To Be Added")
row = write_param(ws, row, "original_value", "Number", "1500.00", "Pre-discount price", "", "Medium", "To Be Added")
row = write_param(ws, row, "items[]", "Array", "(single item)", "Same item params as first_purchase", "", "High", "To Be Added")
row += 1

# ── 16. subscribe ──

SUBSCRIBE_JSON = """// Returning customer subscribes with promo code
window.dataLayer.push({
  'event': 'subscribe',
  'transaction_id': 'pay-b8c9d0e1-...',
  'currency': 'EGP',
  'value': 1875.00,
  'item_type': 'subscription',
  'payment_type': 'fawry',
  'customer_type': 'returning',
  'subscription_duration': 'annual',
  'coupon': 'SUBSCRIBE25',
  'discount': 625.00,
  'original_value': 2500.00,
  'items': [{
    'item_id': 'subscription_annual',
    'item_name': 'TrafficMENA Annual Subscription',
    'item_category': 'Subscription',
    'price': 1875.00,
    'currency': 'EGP',
    'quantity': 1
  }]
});"""

row = write_event_header(ws, row, "subscribe", "subscription purchase confirmed — new or returning customer (Client + Server)")
row = write_param(ws, row, "transaction_id", "String", "pay-a7b8c9d0-...", "Payment UUID", SUBSCRIBE_JSON, "High", "To Be Added")
row = write_param(ws, row, "currency", "String", "EGP", "Dynamic", "", "High", "To Be Added")
row = write_param(ws, row, "value", "Number", "2500.00", "Amount paid", "", "High", "To Be Added")
row = write_param(ws, row, "item_type", "String", "subscription", 'Always "subscription"', "", "High", "To Be Added")
row = write_param(ws, row, "payment_type", "String", "meeza", "Payment method", "", "High", "To Be Added")
row = write_param(ws, row, "customer_type", "String", "new", '"new" (1 prior purchase) / "returning" (2+). Subscription is always paid — free users become "new" after first purchase.', "", "High", "To Be Added")
row = write_param(ws, row, "subscription_duration", "String", "annual", "Subscription period", "", "High", "To Be Added")
row = write_param(ws, row, "coupon", "String", "", "Promo code. Empty if none.", "", "Medium", "To Be Added")
row = write_param(ws, row, "discount", "Number", "0", "Discount amount", "", "Medium", "To Be Added")
row = write_param(ws, row, "original_value", "Number", "2500.00", "Pre-discount price", "", "Medium", "To Be Added")
row = write_param(ws, row, "items[].item_id", "String", "subscription_annual", "Fixed ID for subscription", "", "High", "To Be Added")
row = write_param(ws, row, "items[].item_name", "String", "TrafficMENA Annual Subscription", "Subscription name", "", "High", "To Be Added")
row = write_param(ws, row, "items[].item_category", "String", "Subscription", 'Always "Subscription"', "", "High", "To Be Added")
row = write_param(ws, row, "items[].price", "Number", "2500.00", "Price paid", "", "High", "To Be Added")
row = write_param(ws, row, "items[].currency", "String", "EGP", "Dynamic", "", "High", "To Be Added")
row = write_param(ws, row, "items[].quantity", "Number", "1", "Always 1", "", "High", "To Be Added")
row += 1

# ── 17. click_meeting_link ──

MEETING_JSON = """window.dataLayer.push({
  'event': 'click_meeting_link',
  'item_id': 'evt-a1b2c3d4-...',
  'item_name': 'Advanced Google Ads Workshop',
  'item_category': 'Mastermind',
  'meeting_platform': 'zoom'
});"""

row = write_event_header(ws, row, "click_meeting_link", "user clicks Zoom/Google Meet link to join a live session")
row = write_param(ws, row, "item_id", "String", "evt-a1b2c3d4-...", "Event UUID", MEETING_JSON, "High", "To Be Added")
row = write_param(ws, row, "item_name", "String", "Advanced Google Ads Workshop", "Event title", "", "High", "To Be Added")
row = write_param(ws, row, "item_category", "String", "Mastermind", "Event type", "", "Medium", "To Be Added")
row = write_param(ws, row, "meeting_platform", "String", "zoom", 'Detected from URL: "zoom" / "google_meet" / "teams" / "other"', "", "Medium", "To Be Added")
row += 1

# ── 18. view_content ──

VIEW_CONTENT_JSON = """// Video lesson within a series
window.dataLayer.push({
  'event': 'view_content',
  'content_id': 'lib-d4e5f6a7-...',
  'content_name': 'Google Ads Bidding Strategies',
  'content_type': 'Video',
  'is_premium': true,
  'series_id': 'ser-e5f6a7b8-...',
  'series_name': 'Google Ads Mastery',
  'event_id': ''
});

// Standalone document from library
window.dataLayer.push({
  'event': 'view_content',
  'content_id': 'lib-f6a7b8c9-...',
  'content_name': 'MENA Marketing Trends Report 2026',
  'content_type': 'Document',
  'is_premium': false,
  'series_id': '',
  'series_name': '',
  'event_id': ''
});"""

row = write_event_header(ws, row, "view_content", "user opens a lesson or library asset (video/document/presentation)")
row = write_param(ws, row, "content_id", "String", "lib-d4e5f6a7-...", "Library asset UUID", VIEW_CONTENT_JSON, "High", "To Be Added")
row = write_param(ws, row, "content_name", "String", "Google Ads Bidding Strategies", "Asset title", "", "High", "To Be Added")
row = write_param(ws, row, "content_type", "String", "Video", '"Video" / "Document" / "Presentation"', "", "High", "To Be Added")
row = write_param(ws, row, "is_premium", "Boolean", "true", "Whether asset requires subscription", "", "Medium", "To Be Added")
row = write_param(ws, row, "series_id", "String", "ser-e5f6a7b8-...", "Series UUID if accessed within a series. Empty if standalone.", "", "Medium", "To Be Added")
row = write_param(ws, row, "series_name", "String", "Google Ads Mastery", "Series title if applicable", "", "Medium", "To Be Added")
row = write_param(ws, row, "event_id", "String", "", "Linked event UUID if asset is event-gated. Empty if not.", "", "Low", "To Be Added")
row += 1

# ── 19. download_content ──

DOWNLOAD_JSON = """window.dataLayer.push({
  'event': 'download_content',
  'content_id': 'lib-f6a7b8c9-...',
  'content_name': 'MENA Marketing Trends Report 2026',
  'content_type': 'Document',
  'is_premium': false
});"""

row = write_event_header(ws, row, "download_content", "user downloads a library asset (document, presentation)")
row = write_param(ws, row, "content_id", "String", "lib-f6a7b8c9-...", "Library asset UUID", DOWNLOAD_JSON, "High", "To Be Added")
row = write_param(ws, row, "content_name", "String", "MENA Marketing Trends Report 2026", "Asset title", "", "High", "To Be Added")
row = write_param(ws, row, "content_type", "String", "Document", "File type", "", "High", "To Be Added")
row = write_param(ws, row, "is_premium", "Boolean", "false", "Premium status", "", "Medium", "To Be Added")
row += 1

# ── 20. add_to_calendar ──

CALENDAR_JSON = """window.dataLayer.push({
  'event': 'add_to_calendar',
  'item_id': 'evt-a1b2c3d4-...',
  'item_name': 'Advanced Google Ads Workshop',
  'calendar_type': 'google_calendar'
});"""

row = write_event_header(ws, row, "add_to_calendar", 'user clicks "Add to Google Calendar" or downloads ICS file')
row = write_param(ws, row, "item_id", "String", "evt-a1b2c3d4-...", "Event UUID", CALENDAR_JSON, "High", "To Be Added")
row = write_param(ws, row, "item_name", "String", "Advanced Google Ads Workshop", "Event title", "", "High", "To Be Added")
row = write_param(ws, row, "calendar_type", "String", "google_calendar", '"google_calendar" / "ics_download"', "", "Medium", "To Be Added")
row += 1

# ── 21. calculator_used ──

CALC_JSON = """window.dataLayer.push({
  'event': 'calculator_used',
  'calculator_id': 'roas',
  'calculator_name': 'Return on Ad Spend',
  'calculator_category': 'revenue_value'
});"""

row = write_event_header(ws, row, "calculator_used", "user completes a calculation in any of the 23 calculators")
row = write_param(ws, row, "calculator_id", "String", "roas", "Calculator slug (e.g., cpc, ltv, roas, cvr, mom-growth)", CALC_JSON, "High", "To Be Added")
row = write_param(ws, row, "calculator_name", "String", "Return on Ad Spend", "Display name", "", "High", "To Be Added")
row = write_param(ws, row, "calculator_category", "String", "revenue_value", '"traffic_acquisition" / "conversion" / "revenue_value" / "retention_growth" / "efficiency"', "", "Medium", "To Be Added")
row += 1

# ── 22. cancel_registration ──

CANCEL_JSON = """// Cancel free event
window.dataLayer.push({
  'event': 'cancel_registration',
  'item_id': 'evt-b2c3d4e5-...',
  'item_name': 'SEO Fundamentals Meetup',
  'item_category': 'Meetup',
  'cancellation_type': 'instant',
  'was_paid': false
});

// Request refund for paid event
window.dataLayer.push({
  'event': 'cancel_registration',
  'item_id': 'evt-a1b2c3d4-...',
  'item_name': 'Advanced Google Ads Workshop',
  'item_category': 'Mastermind',
  'cancellation_type': 'refund_request',
  'was_paid': true
});"""

row = write_event_header(ws, row, "cancel_registration", "user cancels event registration or requests a refund (Client + Server)")
row = write_param(ws, row, "item_id", "String", "evt-a1b2c3d4-...", "Event UUID", CANCEL_JSON, "High", "To Be Added")
row = write_param(ws, row, "item_name", "String", "Advanced Google Ads Workshop", "Event title", "", "High", "To Be Added")
row = write_param(ws, row, "item_category", "String", "Mastermind", "Event type", "", "Medium", "To Be Added")
row = write_param(ws, row, "cancellation_type", "String", "refund_request", '"instant" (free) / "refund_request" (paid)', "", "High", "To Be Added")
row = write_param(ws, row, "was_paid", "Boolean", "true", "Whether registration involved a payment", "", "High", "To Be Added")
row += 1

# ── 23. profile_updated ──

PROFILE_JSON = """window.dataLayer.push({
  'event': 'profile_updated',
  'fields_updated': 'phone,primary_goal',
  'profile_completion': 85
});"""

row = write_event_header(ws, row, "profile_updated", "user saves changes to their profile")
row = write_param(ws, row, "fields_updated", "String", "phone,primary_goal", "Comma-separated list of changed fields", PROFILE_JSON, "Medium", "To Be Added")
row = write_param(ws, row, "profile_completion", "Number", "85", "Completion % (0-100). Fields: firstName, lastName, email, phone, goal, challenge", "", "Medium", "To Be Added")


# ════════════════════════════════════════════════════════════════════
# SHEET 3: Signup Steps Reference
# ════════════════════════════════════════════════════════════════════

ws_steps = wb.create_sheet("Signup Steps")
ws_steps.sheet_properties.tabColor = "4472C4"

ws_steps.merge_cells("A1:E1")
ws_steps.cell(row=1, column=1, value="SIGNUP FUNNEL — STEP DEFINITIONS").font = HEADER_FONT
ws_steps.cell(row=1, column=1).fill = HEADER_FILL
for c in range(2, 6):
    ws_steps.cell(row=1, column=c).fill = HEADER_FILL

step_headers = ["step_number", "step_name", "Data Collected", "Component", "Notes"]
for i, h in enumerate(step_headers, 1):
    cell = ws_steps.cell(row=3, column=i, value=h)
    cell.font = COL_HEADER_FONT
    cell.fill = COL_HEADER_FILL

steps = [
    (1, "name_info", "firstName, lastName", "Step1.tsx", "Min 2 chars, letters/spaces/hyphens only"),
    (2, "email_entered", "email address", "Step2.tsx", "Standard email validation"),
    (3, "otp_verified", "OTP code verified", "CheckEmail.tsx", "6-digit code verified against server"),
    (4, "phone_entered", "phone number (WhatsApp)", "Step3.tsx", "Format: +{dialCode}{number}, 7-15 digits"),
    (5, "goal_selected", "primaryGoal", "Step4.tsx", "Single selection from 5 predefined options"),
    (6, "challenge_selected", "primaryChallenge", "Step5.tsx", "Single selection from 5 predefined options"),
]

for i, s in enumerate(steps):
    r = 4 + i
    for j, val in enumerate(s):
        cell = ws_steps.cell(row=r, column=j + 1, value=val)
        cell.font = PARAM_FONT
        cell.border = THIN_BORDER

ws_steps.column_dimensions["A"].width = 15
ws_steps.column_dimensions["B"].width = 22
ws_steps.column_dimensions["C"].width = 30
ws_steps.column_dimensions["D"].width = 18
ws_steps.column_dimensions["E"].width = 50


# ════════════════════════════════════════════════════════════════════
# SHEET 4: Page Types
# ════════════════════════════════════════════════════════════════════

ws_pages = wb.create_sheet("Page Types")
ws_pages.sheet_properties.tabColor = "548235"

ws_pages.merge_cells("A1:C1")
ws_pages.cell(row=1, column=1, value="PAGE TYPE VALUES — for global_variables.page_type").font = HEADER_FONT
ws_pages.cell(row=1, column=1).fill = HEADER_FILL
for c in range(2, 4):
    ws_pages.cell(row=1, column=c).fill = HEADER_FILL

page_headers = ["page_type value", "Route(s)", "Area"]
for i, h in enumerate(page_headers, 1):
    cell = ws_pages.cell(row=3, column=i, value=h)
    cell.font = COL_HEADER_FONT
    cell.fill = COL_HEADER_FILL

pages = [
    ("homepage", "/", "Public"),
    ("about", "/about", "Public"),
    ("event_list", "/meetups", "Public"),
    ("event_detail", "/meetups/:id", "Public"),
    ("track_detail", "/tracks/:id", "Public"),
    ("subscribe_landing", "/subscribe", "Public"),
    ("signin", "/signin", "Auth"),
    ("signup", "/signup/*", "Auth"),
    ("dashboard", "/dashboard", "Dashboard"),
    ("dashboard_events", "/dashboard/meetups", "Dashboard"),
    ("dashboard_library", "/dashboard/library", "Dashboard"),
    ("dashboard_library_detail", "/dashboard/library/:id", "Dashboard"),
    ("dashboard_track_detail", "/dashboard/library/tracks/:id", "Dashboard"),
    ("dashboard_calculators", "/dashboard/calculators", "Dashboard"),
    ("calculator_detail", "/dashboard/calculators/:slug", "Dashboard"),
    ("dashboard_profile", "/dashboard/profile", "Dashboard"),
    ("payment_success", "/payment/success", "Payment"),
    ("payment_failed", "/payment/failed", "Payment"),
    ("payment_pending", "/payment/pending", "Payment"),
    ("thank_you_event", "/thank-you-event/:id", "Payment"),
    ("thank_you_track", "/thank-you-track/:id", "Payment"),
    ("privacy", "/privacy", "Legal"),
    ("terms", "/terms", "Legal"),
]

for i, p in enumerate(pages):
    r = 4 + i
    for j, val in enumerate(p):
        cell = ws_pages.cell(row=r, column=j + 1, value=val)
        cell.font = PARAM_FONT
        cell.border = THIN_BORDER

ws_pages.column_dimensions["A"].width = 28
ws_pages.column_dimensions["B"].width = 38
ws_pages.column_dimensions["C"].width = 14


# ════════════════════════════════════════════════════════════════════
# SHEET 5: Calculator Reference
# ════════════════════════════════════════════════════════════════════

ws_calc = wb.create_sheet("Calculator Reference")
ws_calc.sheet_properties.tabColor = "BF8F00"

ws_calc.merge_cells("A1:C1")
ws_calc.cell(row=1, column=1, value="CALCULATOR SLUGS — for calculator_used.calculator_id").font = HEADER_FONT
ws_calc.cell(row=1, column=1).fill = HEADER_FILL
for c in range(2, 4):
    ws_calc.cell(row=1, column=c).fill = HEADER_FILL

calc_headers = ["calculator_id", "calculator_name", "calculator_category"]
for i, h in enumerate(calc_headers, 1):
    cell = ws_calc.cell(row=3, column=i, value=h)
    cell.font = COL_HEADER_FONT
    cell.fill = COL_HEADER_FILL

calcs = [
    ("cpc", "Cost Per Click", "traffic_acquisition"),
    ("cpm", "Cost Per Mille", "traffic_acquisition"),
    ("cpl", "Cost Per Lead", "traffic_acquisition"),
    ("ctr", "Click-Through Rate", "traffic_acquisition"),
    ("cac", "Customer Acquisition Cost", "traffic_acquisition"),
    ("ncac", "New Customer Acquisition Cost", "traffic_acquisition"),
    ("cac-payback", "CAC Payback Period", "traffic_acquisition"),
    ("cvr", "Conversion Rate", "conversion"),
    ("cart-abandonment", "Cart Abandonment Rate", "conversion"),
    ("checkout-abandonment", "Checkout Abandonment Rate", "conversion"),
    ("lead-to-customer", "Lead to Customer Rate", "conversion"),
    ("repeat-purchase", "Repeat Purchase Rate", "conversion"),
    ("aov", "Average Order Value", "revenue_value"),
    ("ltv", "Customer Lifetime Value", "revenue_value"),
    ("saas-ltv", "SaaS LTV Calculator", "revenue_value"),
    ("ltv-cac", "LTV:CAC Ratio", "revenue_value"),
    ("roas", "Return on Ad Spend", "revenue_value"),
    ("breakeven-roas", "Breakeven ROAS", "revenue_value"),
    ("grr", "Gross Revenue Retention", "retention_growth"),
    ("nrr", "Net Revenue Retention", "retention_growth"),
    ("mom-growth", "Month over Month Growth", "retention_growth"),
    ("mer", "Marketing Efficiency Ratio", "efficiency"),
    ("seo-roi", "SEO ROI Calculator", "efficiency"),
]

for i, c in enumerate(calcs):
    r = 4 + i
    for j, val in enumerate(c):
        cell = ws_calc.cell(row=r, column=j + 1, value=val)
        cell.font = PARAM_FONT
        cell.border = THIN_BORDER

ws_calc.column_dimensions["A"].width = 22
ws_calc.column_dimensions["B"].width = 35
ws_calc.column_dimensions["C"].width = 22


# ════════════════════════════════════════════════════════════════════
# SHEET 6: Implementation Priority
# ════════════════════════════════════════════════════════════════════

ws_pri = wb.create_sheet("Implementation Priority")
ws_pri.sheet_properties.tabColor = "C00000"

ws_pri.merge_cells("A1:D1")
ws_pri.cell(row=1, column=1, value="IMPLEMENTATION PRIORITY MATRIX").font = HEADER_FONT
ws_pri.cell(row=1, column=1).fill = HEADER_FILL
for c in range(2, 5):
    ws_pri.cell(row=1, column=c).fill = HEADER_FILL

phase_headers = ["Phase", "Event", "Rationale", "Status"]
for i, h in enumerate(phase_headers, 1):
    cell = ws_pri.cell(row=3, column=i, value=h)
    cell.font = COL_HEADER_FONT
    cell.fill = COL_HEADER_FILL

phases = [
    ("Phase 1: Core Funnels", "global_variables", "Foundation for all other events", "Not Started"),
    ("Phase 1: Core Funnels", "login_start + login", "Auth funnel measurement", "Not Started"),
    ("Phase 1: Core Funnels", "sign_up_step + sign_up", "Acquisition funnel", "Not Started"),
    ("Phase 1: Core Funnels", "view_item", "Content interest", "Not Started"),
    ("Phase 1: Core Funnels", "begin_checkout", "Payment funnel top", "Not Started"),
    ("Phase 1: Core Funnels", "select_payment_method", "Payment funnel middle", "Not Started"),
    ("Phase 1: Core Funnels", "first_purchase", "New customer conversion (ads)", "Not Started"),
    ("Phase 1: Core Funnels", "purchase", "Returning customer conversion (ads)", "Not Started"),
    ("Phase 1: Core Funnels", "subscribe", "Subscription conversion (ads)", "Not Started"),
    ("Phase 2: Discovery", "view_item_list", "Content discovery", "Not Started"),
    ("Phase 2: Discovery", "select_item", "Click-through analysis", "Not Started"),
    ("Phase 2: Discovery", "event_registration", "Free conversion tracking", "Not Started"),
    ("Phase 2: Discovery", "track_booking", "Free conversion tracking", "Not Started"),
    ("Phase 2: Discovery", "click_meeting_link", "Attendance tracking", "Not Started"),
    ("Phase 2: Discovery", "view_content", "Content consumption", "Not Started"),
    ("Phase 3: Optimization", "apply_promo_code", "Promotion effectiveness", "Not Started"),
    ("Phase 3: Optimization", "cancel_registration", "Churn analysis", "Not Started"),
    ("Phase 3: Optimization", "download_content", "Content value", "Not Started"),
    ("Phase 3: Optimization", "add_to_calendar", "Attendance intent", "Not Started"),
    ("Phase 3: Optimization", "calculator_used", "Tool engagement", "Not Started"),
    ("Phase 3: Optimization", "profile_updated", "Engagement depth", "Not Started"),
]

for i, p in enumerate(phases):
    r = 4 + i
    for j, val in enumerate(p):
        cell = ws_pri.cell(row=r, column=j + 1, value=val)
        cell.font = PARAM_FONT
        cell.border = THIN_BORDER
    if "Phase 1" in p[0]:
        ws_pri.cell(row=r, column=1).fill = HIGH_FILL
    elif "Phase 2" in p[0]:
        ws_pri.cell(row=r, column=1).fill = MEDIUM_FILL
    else:
        ws_pri.cell(row=r, column=1).fill = LOW_FILL

ws_pri.column_dimensions["A"].width = 28
ws_pri.column_dimensions["B"].width = 28
ws_pri.column_dimensions["C"].width = 40
ws_pri.column_dimensions["D"].width = 14


# ════════════════════════════════════════════════════════════════════
# SHEET 7: Server-Side Events
# ════════════════════════════════════════════════════════════════════

ws_srv = wb.create_sheet("Server-Side Events")
ws_srv.sheet_properties.tabColor = "7030A0"

ws_srv.merge_cells("A1:E1")
ws_srv.cell(row=1, column=1, value="SERVER-SIDE EVENTS — GA4 Measurement Protocol").font = HEADER_FONT
ws_srv.cell(row=1, column=1).fill = HEADER_FILL
for c in range(2, 6):
    ws_srv.cell(row=1, column=c).fill = HEADER_FILL

srv_headers = ["Event", "Server Trigger Point", "Why Server-Side", "Deduplication Key", "Notes"]
for i, h in enumerate(srv_headers, 1):
    cell = ws_srv.cell(row=3, column=i, value=h)
    cell.font = COL_HEADER_FONT
    cell.fill = COL_HEADER_FILL

srv_events = [
    ("sign_up", "After profile record insert", "User may close browser after last step", "dedup_id = signup-{userId}", "Fire after both user + profile records confirmed"),
    ("event_registration", "After eventAttendees insert", "Ensures registration is recorded", "dedup_id = reg-{eventId}-{userId}", "Only for free events"),
    ("track_booking", "After trackBookings insert", "Ensures booking is recorded", "dedup_id = book-{trackId}-{userId}", "Only for free tracks"),
    ("first_purchase", "Fawaterk webhook, after fulfillment", "Payment confirmation from gateway", "transaction_id = payments.id", "Authoritative purchase record. Check prior purchases count = 0."),
    ("purchase", "Fawaterk webhook, after fulfillment", "Payment confirmation from gateway", "transaction_id = payments.id", "Authoritative purchase record. Check prior purchases count > 0."),
    ("subscribe", "Fawaterk webhook, after subscription created", "Payment confirmation from gateway", "transaction_id = payments.id", "Fire after subscription record is created."),
    ("cancel_registration", "After eventAttendees status update", "Server-initiated cancellations", "dedup_id = cancel-{eventId}-{userId}", "Status → cancelled or refund_requested"),
]

for i, s in enumerate(srv_events):
    r = 4 + i
    for j, val in enumerate(s):
        cell = ws_srv.cell(row=r, column=j + 1, value=val)
        cell.font = PARAM_FONT
        cell.border = THIN_BORDER
        cell.alignment = WRAP_ALIGN

ws_srv.column_dimensions["A"].width = 22
ws_srv.column_dimensions["B"].width = 35
ws_srv.column_dimensions["C"].width = 38
ws_srv.column_dimensions["D"].width = 35
ws_srv.column_dimensions["E"].width = 50


# ════════════════════════════════════════════════════════════════════
# SHEET 8: GA4 Custom Dimensions
# ════════════════════════════════════════════════════════════════════

ws_dims = wb.create_sheet("GA4 Custom Dimensions")
ws_dims.sheet_properties.tabColor = "00B050"

ws_dims.merge_cells("A1:C1")
ws_dims.cell(row=1, column=1, value="GA4 CUSTOM DIMENSIONS — Register in GA4 Admin").font = HEADER_FONT
ws_dims.cell(row=1, column=1).fill = HEADER_FILL
for c in range(2, 4):
    ws_dims.cell(row=1, column=c).fill = HEADER_FILL

dim_headers = ["Parameter Name", "Scope", "Type"]
for i, h in enumerate(dim_headers, 1):
    cell = ws_dims.cell(row=3, column=i, value=h)
    cell.font = COL_HEADER_FONT
    cell.fill = COL_HEADER_FILL

dims = [
    ("user_role", "User", "Text"),
    ("subscription_status", "User", "Text"),
    ("customer_type", "User", "Text"),
    ("item_type", "Event", "Text"),
    ("payment_type", "Event", "Text"),
    ("item_category", "Event", "Text"),
    ("registration_type", "Event", "Text"),
    ("booking_type", "Event", "Text"),
    ("content_type", "Event", "Text"),
    ("calculator_category", "Event", "Text"),
    ("cancellation_type", "Event", "Text"),
    ("is_online", "Event", "Boolean"),
    ("is_premium", "Event", "Boolean"),
    ("step_name", "Event", "Text"),
]

for i, d in enumerate(dims):
    r = 4 + i
    for j, val in enumerate(d):
        cell = ws_dims.cell(row=r, column=j + 1, value=val)
        cell.font = PARAM_FONT
        cell.border = THIN_BORDER

ws_dims.column_dimensions["A"].width = 24
ws_dims.column_dimensions["B"].width = 12
ws_dims.column_dimensions["C"].width = 12


# ════════════════════════════════════════════════════════════════════
# SHEET 9: Ad Platform Conversions
# ════════════════════════════════════════════════════════════════════

ws_ads = wb.create_sheet("Ad Platform Conversions")
ws_ads.sheet_properties.tabColor = "FF6600"

ws_ads.merge_cells("A1:D1")
ws_ads.cell(row=1, column=1, value="AD PLATFORM CONVERSION MAPPING").font = HEADER_FONT
ws_ads.cell(row=1, column=1).fill = HEADER_FILL
for c in range(2, 5):
    ws_ads.cell(row=1, column=c).fill = HEADER_FILL

# Google Ads section
ws_ads.merge_cells("A3:D3")
ws_ads.cell(row=3, column=1, value="Google Ads Conversion Actions").font = Font(name="Calibri", bold=True, size=12)

ads_headers = ["Conversion Action", "GA4 Event", "Bid Strategy", "Notes"]
for i, h in enumerate(ads_headers, 1):
    cell = ws_ads.cell(row=4, column=i, value=h)
    cell.font = COL_HEADER_FONT
    cell.fill = COL_HEADER_FILL

google_ads = [
    ("New Customer Purchase", "first_purchase", "Max conversion value (new customer acquisition)", "Primary acquisition goal"),
    ("Repeat Purchase", "purchase", "Max conversion value (retention)", "Secondary goal"),
    ("Subscription", "subscribe", "Max conversion value (high LTV)", "Highest-value conversion"),
    ("Sign Up", "sign_up", "Max conversions (top of funnel)", "Awareness/engagement campaigns"),
    ("Free Registration", "event_registration", "Max conversions (engagement)", "Community growth campaigns"),
]

for i, a in enumerate(google_ads):
    r = 5 + i
    for j, val in enumerate(a):
        cell = ws_ads.cell(row=r, column=j + 1, value=val)
        cell.font = PARAM_FONT
        cell.border = THIN_BORDER

# Meta/TikTok section
ws_ads.merge_cells("A12:D12")
ws_ads.cell(row=12, column=1, value="Meta / TikTok Event Mapping").font = Font(name="Calibri", bold=True, size=12)

meta_headers = ["Platform Event", "DataLayer Event", "Notes", ""]
for i, h in enumerate(meta_headers, 1):
    cell = ws_ads.cell(row=13, column=i, value=h)
    cell.font = COL_HEADER_FONT
    cell.fill = COL_HEADER_FILL

meta_events = [
    ("CompleteRegistration", "sign_up", "Maps to signup completion", ""),
    ("Purchase", "first_purchase / purchase", "Both map to Meta Purchase", ""),
    ("Subscribe", "subscribe", "Maps to Meta Subscribe", ""),
    ("ViewContent", "view_item", "Maps to content view", ""),
    ("InitiateCheckout", "begin_checkout", "Maps to checkout start", ""),
    ("AddPaymentInfo", "select_payment_method", "Maps to payment method selection", ""),
]

for i, m in enumerate(meta_events):
    r = 14 + i
    for j, val in enumerate(m):
        cell = ws_ads.cell(row=r, column=j + 1, value=val)
        cell.font = PARAM_FONT
        cell.border = THIN_BORDER

ws_ads.column_dimensions["A"].width = 28
ws_ads.column_dimensions["B"].width = 28
ws_ads.column_dimensions["C"].width = 45
ws_ads.column_dimensions["D"].width = 20


# ════════════════════════════════════════════════════════════════════
# Save
# ════════════════════════════════════════════════════════════════════

output_path = "docs/TrafficMENA - Events Tracking Data Model.xlsx"
wb.save(output_path)
print(f"Excel file saved to: {output_path}")
